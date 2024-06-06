import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import StaleElementReferenceException
from datetime import datetime, timedelta
import time
import tkinter as tk
import requests
import subprocess
from tkinter import filedialog, messagebox
import threading
import os
import sys
from openpyxl.styles import Font
from tkinter import ttk
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from urllib.parse import urljoin



class App:
    def __init__(self, root):
        self.root = root
        self.root.title("PHẦN MỀM TỰ ĐỘNG HÓA THAO TÁC HỆ THỐNG MỘT CỬA ĐIỆN TỬ")
        self.root.geometry("700x550")

        self.file_path = ""
        self.stop_flag = False

        # Thiết lập font chữ tổng thể
        self.default_font = ("Helvetica", 13)
        self.default_font2 = ("Helvetica", 10)
        self.button_font = ("Helvetica", 13, "bold")
        self.button_font2 = ("Helvetica", 10, "bold")
        self.title_font = ("Helvetica", 16, "bold")
        self.signature_font = ("Helvetica",9)
        

        # Tạo nút "Check Update"
        self.check_update_button = tk.Button(root, text="Check Update", command=self.check_for_update)
        self.check_update_button.pack(anchor='se', padx=5, pady=5)
        
        # Tiêu đề
        self.title_label = tk.Label(root, text="PHẦN MỀM TỰ ĐỘNG HÓA THAO TÁC\nHỆ THỐNG MỘT CỬA ĐIỆN TỬ", fg="red", font=self.title_font)
        self.title_label.pack(pady=10)
        

        # Khung chứa phần chọn file
        self.file_frame = tk.Frame(root)
        self.file_frame.pack(pady=10, padx=10, fill=tk.X)

        self.file_entry = tk.Entry(self.file_frame, width=50, font=self.default_font)
        self.file_entry.pack(side=tk.LEFT, padx=5, expand=True, fill=tk.X)

        self.open_button = tk.Button(self.file_frame, text="Open", command=self.open_file, font=self.default_font, fg="blue")
        self.open_button.pack(side=tk.LEFT, padx=5)
        
        

        # Khung chứa các checkbutton
        self.checkbox_frame = tk.Frame(root)
        self.checkbox_frame.pack(pady=10, padx=10, anchor='w')

        self.checkbox_dongboTTC = tk.IntVar() #đồng bộ TTC
        self.checkbox_dongboDVC = tk.IntVar() #đồng bộ DVC
        self.checkbox_dongbolv = tk.IntVar() #đồng bộ lĩnh vực
        self.checkbox_offnamsau = tk.IntVar() #Cấu hình nghỉ lễ năm sau
        self.checkbox_copysovb = tk.IntVar() #sao chép sổ vb
        self.holiday_year_choice = tk.StringVar(value="Năm sau") # Thêm biến cho drop-down
        self.checkbox_quytrinh = tk.IntVar()
        
        self.checkbox4 = tk.Checkbutton(self.checkbox_frame, text="Đồng bộ Lĩnh Vực", variable=self.checkbox_dongbolv, font=self.default_font)
        self.checkbox4.pack(anchor='w')
        
        self.checkbox1 = tk.Checkbutton(self.checkbox_frame, text="Đồng bộ Thủ Tục Chung", variable=self.checkbox_dongboTTC, font=self.default_font)
        self.checkbox1.pack(anchor='w')

        self.checkbox2 = tk.Checkbutton(self.checkbox_frame, text="Đồng bộ Dịch Vụ Công", variable=self.checkbox_dongboDVC, font=self.default_font)
        self.checkbox2.pack(anchor='w')

        # Tạo frame con để chứa Checkbutton và drop-down menu cho "Cấu hình Ngày Nghỉ Lễ"
        self.date_config_frame = tk.Frame(self.checkbox_frame)
        self.date_config_frame.pack(anchor='w')

        self.checkbox_offnamsau_var = tk.IntVar()
        self.checkbox_offnamsau = tk.Checkbutton(self.date_config_frame, text="Cấu hình Ngày Nghỉ Lễ", variable=self.checkbox_offnamsau_var, font=self.default_font)
        self.checkbox_offnamsau.pack(side=tk.LEFT)

        self.year_menu = ttk.Combobox(self.date_config_frame, textvariable=self.holiday_year_choice, values=["Năm sau", "Năm nay"], font=self.default_font, state='readonly')
        self.year_menu.pack(side=tk.LEFT)
        
        # Checkbox "Sao chép Sổ Văn Bản"
        self.checkbox_copysovb_var = tk.IntVar()
        self.checkbox_copysovb = tk.Checkbutton(self.checkbox_frame, text="Sao chép Sổ Văn Bản", variable=self.checkbox_copysovb_var, font=self.default_font)
        self.checkbox_copysovb.pack(anchor='w')

        # Khung chứa checkbox "Cấu hình quy trình tự động" và nút "Đính kèm"
        self.checkbox_quytrinh_frame = tk.Frame(self.checkbox_frame)
        self.checkbox_quytrinh_frame.pack(anchor='w')

        self.checkbox_quytrinh_var = tk.IntVar()
        self.checkbox_quytrinh = tk.Checkbutton(self.checkbox_quytrinh_frame, text="Cấu hình Quy Trình tự động", variable=self.checkbox_quytrinh_var, font=self.default_font)
        self.checkbox_quytrinh.pack(side=tk.LEFT)

        self.attach_button = tk.Button(self.checkbox_quytrinh_frame, text="Chọn quy trình", command=self.attach_file, font=self.default_font)
        self.attach_button.pack(side=tk.LEFT, padx=5)

        # Label để hiển thị đường dẫn tệp đã đính kèm
        self.attached_file_label = tk.Label(self.checkbox_quytrinh_frame, text="", font=self.default_font)
        self.attached_file_label.pack(side=tk.LEFT)

        # Biến lưu đường dẫn tệp đính kèm
        self.attached_file_path = None
        
        # Khung chứa các nút tải file
        self.download_frame = tk.Frame(root)
        self.download_frame.pack(pady=5)

        # Nút "Tải file mẫu"
        self.download_button = tk.Button(self.download_frame, text="Tải file mẫu", command=self.download_sample_file, font=self.button_font, fg="blue")
        self.download_button.pack(side=tk.LEFT, padx=5)

        # Nút "Cấu hình quy trình"
        self.download_button_quytrinh = tk.Button(self.download_frame, text="Cấu hình quy trình", command=self.open_quy_trinh_window, font=self.button_font, fg="brown")
        self.download_button_quytrinh.pack(side=tk.LEFT, padx=5)

        
        # Các nút điều khiển
        self.control_frame = tk.Frame(root)
        self.control_frame.pack(pady=10)

        self.start_button = tk.Button(self.control_frame, text="START", command=self.start_thread, font=self.button_font, fg="green")
        self.start_button.grid(row=0, column=0, padx=5)

        self.stop_button = tk.Button(self.control_frame, text="STOP", command=self.stop_automation, state=tk.DISABLED, font=self.button_font, fg="red")
        self.stop_button.grid(row=0, column=1, padx=5)


        # Chữ ký
        self.signature_label = tk.Label(root, text="Phòng Ứng dụng CNTT - Trung tâm Công nghệ thông tin và Truyền thông",fg="purple", font=self.signature_font, anchor='e')
        self.signature_label.pack(side=tk.BOTTOM, padx=10, pady=10, anchor='se')
        
    
    def read_quy_trinh_file(self):
        try:
            workbook = load_workbook(self.attached_file_path)
            sheet = workbook["QuyTrinh"]
            self.quy_trinh_data = pd.DataFrame(sheet.values)
            messagebox.showinfo("Thông báo", "Đọc file quy trình thành công!")
        except Exception as e:
            messagebox.showerror("Lỗi", f"Lỗi khi đọc file quy trình: {str(e)}")
            self.quy_trinh_data = None
        
    
    def check_for_update(self):
        # URL của tệp cập nhật trên GitHub
        repo_owner = "duy97ct"
        repo_name = "motcua_auto"
        file_path = "Motcua_auto.exe"  # Đường dẫn tệp trên GitHub

        # Đường dẫn lưu tệp trên hệ thống (cùng thư mục với file .exe)
        if getattr(sys, 'frozen', False):
            application_path = os.path.dirname(sys.executable)
        else:
            application_path = os.path.dirname(os.path.abspath(__file__))
            
        save_path = os.path.join(application_path, "Motcua_auto.exe")
        
        # URL trực tiếp tới tệp .exe trên GitHub
        url = f"https://github.com/{repo_owner}/{repo_name}/raw/main/{file_path}"
        
        response = requests.get(url, stream=True)
        
        if response.status_code == 200:
            if messagebox.askyesno("Cập nhật", "Có bản cập nhật mới. Bạn có muốn tải xuống và cài đặt không?"):
                with open(save_path, 'wb') as file:
                    for chunk in response.iter_content(chunk_size=8192):
                        file.write(chunk)
                messagebox.showinfo("Thông báo", "Cập nhật thành công. Vui lòng khởi động lại ứng dụng.")
                self.root.destroy()
            else:
                messagebox.showinfo("Thông báo", "Đã hủy cập nhật.")
        else:
            messagebox.showerror("Lỗi", "Không thể tải tệp cập nhật. Vui lòng thử lại sau.")


    def download_and_replace_update(self):
        try:
            # URL tải về tệp cập nhật .exe
            update_url = "https://github.com/duy97ct/motcua_auto/raw/main/Motcua_auto.exe"
            
            # Đường dẫn tới tệp .exe hiện tại
            current_exe_path = os.path.abspath(__file__)

            # Tạo một tên tạm cho tệp tải về
            temp_exe_path = current_exe_path + ".temp"

            # Tải xuống tệp cập nhật
            response = requests.get(update_url, stream=True)
            response.raise_for_status()

            # Lưu tệp cập nhật tạm thời
            with open(temp_exe_path, "wb") as f:
                for chunk in response.iter_content(chunk_size=8192):
                    f.write(chunk)

            # Thay thế tệp hiện tại bằng tệp cập nhật
            os.replace(temp_exe_path, current_exe_path)

            # Hiển thị thông báo và thoát ứng dụng
            messagebox.showinfo("Cập nhật", "Đã tải xuống cập nhật. Vui lòng khởi động lại ứng dụng.")
            self.root.destroy()
            
            # Tùy chọn: Tự động khởi động lại ứng dụng
            subprocess.Popen([current_exe_path])

        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể tải cập nhật: {e}")   

    def open_file(self):
        self.file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if self.file_path:
            self.file_entry.delete(0, tk.END)
            self.file_entry.insert(0, self.file_path)

    def start_thread(self):
        self.stop_flag = False
        self.stop_button.config(state=tk.NORMAL)
        thread = threading.Thread(target=self.start_automation)
        thread.start()

    def stop_automation(self):
        self.stop_flag = True

    def start_automation(self):
    
        
        # Xác định đường dẫn của ChromeDriver
        if getattr(sys, 'frozen', False):
            chromedriver_path = os.path.join(sys._MEIPASS, "chromedriver.exe")
        else:
            chromedriver_path = "chromedriver.exe"

        
        # # Khởi tạo trình duyệt
        # # driver = webdriver.Chrome(chromedriver_path) #THƯ MỤC GỐC
        
        # Khởi tạo ChromeOptions
        chrome_options = webdriver.ChromeOptions()

        # Vô hiệu hóa chế độ ẩn danh
        chrome_options.add_argument("--disable-blink-features=AutomationControlled")

        # Đường dẫn đến thư mục lưu trữ dữ liệu cá nhân của Chrome
        chrome_options.add_argument("user-data-dir=C:\\chromedriver")

        # Khởi tạo trình duyệt với ChromeOptions và đường dẫn của ChromeDriver
        driver = webdriver.Chrome(executable_path=chromedriver_path, options=chrome_options)
        
        df = pd.read_excel(self.file_path)
        
        try:
            
            for index, row in df.iterrows():
                print(f"Processing row {index+1}")
                if self.stop_flag:
                    print("Stop flag set. Exiting loop.")
                    break

                url = row['URL']
                next_year = datetime.now().year +1
                now_year = datetime.now().year
                value1 = row['admin']
                value2 = row['pass']

                # Đường dẫn đồng bộ TTC
                value3 = url + "/group/guest/danh-muc?p_p_id=DanhMuc_WAR_ctonegateportlet&p_p_lifecycle=1&p_p_state=normal&p_p_mode=view&_DanhMuc_WAR_ctonegateportlet_javax.portlet.action=viewDanhMucThuTucChung"
                
                # Đường dẫn đồng bộ DVC
                value4 = url + "/group/guest/danh-muc?p_auth=LxgJHKS6&p_p_id=DanhMuc_WAR_ctonegateportlet&p_p_lifecycle=1&p_p_state=normal&p_p_mode=view&_DanhMuc_WAR_ctonegateportlet_javax.portlet.action=viewDanhMucThuTuc"
                
                # Đường dẫn lịch làm việc
                value5 = url + "/group/guest/lich-lam-viec?p_p_id=quanlylichlamviec_WAR_ctonegatecoreportlet&p_p_lifecycle=0&p_p_state=normal&p_p_mode=view&p_p_col_id=column-2&p_p_col_count=1&_quanlylichlamviec_WAR_ctonegatecoreportlet_isEdit=false&_quanlylichlamviec_WAR_ctonegatecoreportlet_jspPage=%2Fhtml%2Flichlamviec%2Fform_schedule.jsp"
                
                # Đường dẫn đồng bộ lĩnh vực
                value6 = url + "/group/guest/danh-muc?p_auth=LxgJHKS6&p_p_id=DanhMuc_WAR_ctonegateportlet&p_p_lifecycle=1&p_p_state=normal&p_p_mode=view&_DanhMuc_WAR_ctonegateportlet_javax.portlet.action=viewDanhMucLinhVucMotCua"

                # Đường dẫn sao chép sổ văn bản
                value7 = url + "/group/guest/danh-muc?p_p_id=DanhMuc_WAR_ctonegateportlet&p_p_lifecycle=1&p_p_state=normal&p_p_mode=view&p_p_col_id=column-2&p_p_col_count=1&_DanhMuc_WAR_ctonegateportlet_javax.portlet.action=saoChepDanhMucSoVanBan"

                # Cấu hình nghỉ dương lịch
                duonglich_from = row['off_duonglich_from'] #Nghỉ Dương Lịch từ
                if isinstance(duonglich_from, pd.Timestamp):
                    duonglich_from = duonglich_from.strftime('%d/%m/%Y')
                
                duonglich_to = row['off_duonglich_to'] #Nghỉ Dương Lịch đến
                if isinstance(duonglich_to, pd.Timestamp):
                    duonglich_to = duonglich_to.strftime('%d/%m/%Y')
                
                
                # Cấu hình nghỉ nguyên đán
                nguyendan_from = row['off_nguyendan_from'] #Nghỉ Tết Nguyên Đán từ
                if isinstance(nguyendan_from, pd.Timestamp):
                    nguyendan_from = nguyendan_from.strftime('%d/%m/%Y')
                
                nguyendan_to = row['off_nguyendan_to'] #Nghỉ Tết Nguyên Đán đến
                if isinstance(nguyendan_to, pd.Timestamp):
                    nguyendan_to = nguyendan_to.strftime('%d/%m/%Y')

                # Cấu hình nghỉ Giỗ Tổ Hùng Vương
                gioto_from = row['off_gioto_from'] #nghỉ Giỗ Tổ Hùng Vương từ
                if isinstance(gioto_from, pd.Timestamp):
                    gioto_from = gioto_from.strftime('%d/%m/%Y')
                
                gioto_to = row['off_gioto_to'] #nghỉ Giỗ Tổ Hùng Vương đến
                if isinstance(gioto_to, pd.Timestamp):
                    gioto_to = gioto_to.strftime('%d/%m/%Y')

                #Cấu hình nghỉ 30/4 và 1/5                                                                                       
                giaiphong_from = row['off_30/4_va_1/5_from'] #nghỉ 30/4 và 1/5 từ
                if isinstance(giaiphong_from, pd.Timestamp):
                    giaiphong_from = giaiphong_from.strftime('%d/%m/%Y')
                
                giaiphong_to = row['off_30/4_va_1/5_to'] #nghỉ 30/4 và 1/5 đến
                if isinstance(giaiphong_to, pd.Timestamp):
                    giaiphong_to = giaiphong_to.strftime('%d/%m/%Y')

                #Cấu hình Nghỉ lễ 2/9
                quockhanh_from = row['off_2/9_from'] #nghỉ Quốc khánh từ
                if isinstance(quockhanh_from, pd.Timestamp):
                    quockhanh_from = quockhanh_from.strftime('%d/%m/%Y')
                
                quockhanh_to = row['off_2/9_to'] #nghỉ Quốc khánh đến
                if isinstance(quockhanh_to, pd.Timestamp):
                    quockhanh_to = quockhanh_to.strftime('%d/%m/%Y')

                driver.get(url)
                login = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.ID, "_58_login")))
                login.send_keys(value1)

                logout = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.ID, "_58_password")))
                logout.send_keys(value2)

                logout.send_keys(Keys.ENTER)
                
                # Đảm bảo người dùng đã đính kèm file quy trình nếu checkbox được chọn
                if self.checkbox_quytrinh_var.get() and not self.attached_file_path:
                    messagebox.showwarning("Cảnh báo", "Vui lòng chọn file quy trình.")
                    return
                
                # Sử dụng dữ liệu từ df_quytrinh để thực hiện các bước cụ thể nếu checkbox được chọn
                if self.checkbox_quytrinh_var.get():
                    for qt_index, qt_row in self.df_quytrinh.iterrows():
                        print(f"Processing QuyTrinh row {qt_index + 1}")
                        if self.stop_flag:
                            print("Stop flag set. Exiting loop.")
                            break
                        
                        # Các dữ liệu được lấy ra từ file Excel
                        tenqt_value = qt_row['Tên quy trình']  # Tên quy trình
                        tenbidanh_value = qt_row['Bí danh']  # Tên bí danh
                        tenform_value = qt_row['Tên Form']
                        action_value = qt_row['Mã Action']
                        time_value = qt_row['Thời gian']
                        user_group_value = qt_row['Nhóm người dùng']
                        # phongban_value = qt_row['Phòng ban']
                        #Cấu hình quy trình
                        qt_url = url + "/group/guest/quan-tri-quy-trinh?p_p_id=quanlyquytrinh_WAR_ctonegatecoreportlet&p_p_lifecycle=0&p_p_state=normal&p_p_mode=view&p_p_col_id=column-2&p_p_col_count=1&_quanlyquytrinh_WAR_ctonegatecoreportlet_tabs1=Quản+lý+quy+trình&_quanlyquytrinh_WAR_ctonegatecoreportlet_jspPage=%2Fhtml%2Fqlquytrinh%2Fqlqtquytrinh%2Fqlqtquytrinh_add.jsp"
                        #Danh sách Form
                        add_qt_url = url +"/group/guest/quan-tri-quy-trinh?p_p_id=quanlyquytrinh_WAR_ctonegatecoreportlet&p_p_lifecycle=0&p_p_state=normal&p_p_mode=view&p_p_col_id=column-2&p_p_col_count=1&_quanlyquytrinh_WAR_ctonegatecoreportlet_tabs1=Quản+lý+form&_quanlyquytrinh_WAR_ctonegatecoreportlet_jspPage=%2Fhtml%2Fqlquytrinh%2Fqlqtform%2Fqlqtform_add.jsp"
                        
                        #Mở giao diện cấu hình quy trình
                        driver.get(qt_url)
                        bidanhqt = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.ID, "_quanlyquytrinh_WAR_ctonegatecoreportlet_qtQuyTrinhAlias")))
                        bidanhqt.send_keys(tenbidanh_value)
                        
                        tenqt = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.ID, "_quanlyquytrinh_WAR_ctonegatecoreportlet_qtQuyTrinhName")))
                        tenqt.send_keys(tenqt_value)
                        
                        bidanhsave = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.ID, "_quanlyquytrinh_WAR_ctonegatecoreportlet_add")))
                        bidanhsave.click()
                        
                        # tenqt.send_keys(Keys.ENTER) #Nhấn Enter để lưu

                        
                        #Mở giao diện Cấu hình Form
                        driver.get(add_qt_url)
                        
                        #Gắn vào tên form
                        gantenform = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.ID, "_quanlyquytrinh_WAR_ctonegatecoreportlet_TenForm")))
                        gantenform.send_keys(tenform_value)
                        
                        #chọn quy trình xử lý
                        chonqt = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.CLASS_NAME, "select2-selection__rendered")))
                        chonqt.click()
                        
                        gantenqt = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.CLASS_NAME, "select2-search__field")))
                        gantenqt.send_keys(tenqt_value)
                        time.sleep(1)
                        gantenqt.send_keys(Keys.ENTER) #Nhấn Enter để lưu
                        
                         
                        
                        #Cấu hình thời gian(Ngày)
                        timexuly = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.ID, "_quanlyquytrinh_WAR_ctonegatecoreportlet_HanXuLyTheoNgay")))
                        timexuly.clear()  # Xóa dữ liệu hiện có
                        timexuly.send_keys(time_value)
                        
                        #Cấu hình mã action
                        action = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.ID, "_quanlyquytrinh_WAR_ctonegatecoreportlet_MaAction")))
                        action.send_keys(action_value)
                        time.sleep(1)
                        
                        # # Tìm dòng có chứa văn bản từ cột "Nhóm người dùng" và check vào checkbox bên cạnh
                        # Chờ đợi cho bảng hiển thị
                        table = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, "tableqtquyen")))

                        # Tìm tất cả các dòng trong bảng
                        rows = table.find_elements(By.CLASS_NAME, "rowChiTiet")

                        # Duyệt qua từng dòng và chọn checkbox dựa trên nội dung của cột đầu tiên
                        for row in rows:
                            # Lấy văn bản trong thẻ td đầu tiên
                            role_name = row.find_element(By.TAG_NAME, "td").text.strip()
                            # Kiểm tra xem role_name có chứa giá trị bạn muốn không
                            if user_group_value in role_name:
                                # Tìm checkbox
                                checkbox = row.find_element(By.CLASS_NAME, "roleInput")
                                # Kiểm tra nếu không được chọn thì kích vào nó
                                if not checkbox.is_selected():
                                    checkbox.click()
                                break  # Dừng vòng lặp sau khi đã tìm thấy và kích vào checkbox  

                        save = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.ID, "_quanlyquytrinh_WAR_ctonegatecoreportlet_add")))
                        save.click()
                        
                        
                    # Đọc dữ liệu từ sheet "LuanChuyen" trong file quy trình
                    self.df_luanchuyen = pd.read_excel(self.attached_file_path, sheet_name='LuanChuyen', header=4)
                    
                    
                    # Thao tác với dữ liệu từ sheet "LuanChuyen"                    
                    for lc_index, lc_row in self.df_luanchuyen.iterrows():
                        print(f"Processing QuyTrinh row {lc_index + 1}")
                        if self.stop_flag:
                            print("Stop flag set. Exiting loop.")
                            break
                        
                        #Lấy dữ liệu từ sheet LuanChuyen
                        tu_form = lc_row['Từ Form']
                        den_form = lc_row['Đến Form']
                        den_form2 = lc_row['Đến Form 2']
                        den_form3 = lc_row['Đến Form 3']
                        
                        
                        #Cấu hình luân chuyển
                        lc_url = url + "/group/guest/quan-tri-quy-trinh?p_p_id=quanlyquytrinh_WAR_ctonegatecoreportlet&p_p_lifecycle=0&p_p_state=normal&p_p_mode=view&p_p_col_id=column-2&p_p_col_count=1&_quanlyquytrinh_WAR_ctonegatecoreportlet_tabs1=Quản+lý+luân+chuyển&_quanlyquytrinh_WAR_ctonegatecoreportlet_jspPage=%2Fhtml%2Fqlquytrinh%2Fqlqtluanchuyen%2Fqlqtluanchuyen_add.jsp&_quanlyquytrinh_WAR_ctonegatecoreportlet_qtQTQuyTrinhID=-1"
                        
                        #Mở giao diện cấu hình quy trình
                        driver.get(lc_url)
                        
                        #chọn quy trình xử lý
                        chonqt_lc = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.CLASS_NAME, "select2-selection__rendered")))
                        chonqt_lc.click()
                        
                        gantenqt_lc = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.CLASS_NAME, "select2-search__field")))
                        gantenqt_lc.send_keys(tenqt_value)
                        
                        gantenqt_lc.send_keys(Keys.ENTER) #Nhấn Enter để lưu
                        time.sleep(2)
                        
                        #Cấu hình Từ form
                        from_form = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.ID, "_quanlyquytrinh_WAR_ctonegatecoreportlet_qtLC_fromformID")))
                        from_form.send_keys(tu_form)
                        time.sleep(1)
                        
                        #Cấu hình Đến form
                        to_form = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.ID, "_quanlyquytrinh_WAR_ctonegatecoreportlet_qtLC_toformID_01")))
                        to_form.send_keys(den_form)
                        time.sleep(1)
                        
                        #Cấu hình Đến form2
                        to_form2 = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.ID, "_quanlyquytrinh_WAR_ctonegatecoreportlet_qtLC_toformID_02")))
                        to_form2.send_keys(den_form2)
                        time.sleep(1)
                        
                        #Cấu hình Đến form3
                        to_form3 = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.ID, "_quanlyquytrinh_WAR_ctonegatecoreportlet_qtLC_toformID_03")))
                        to_form3.send_keys(den_form3)
                        time.sleep(1)
                
                        save_lc = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.ID, "_quanlyquytrinh_WAR_ctonegatecoreportlet_add")))
                        save_lc.click()
                        
                    #Cấu hình gắn vào TTHC
                    url_dvc = url + "/group/guest/danh-muc?p_p_id=DanhMuc_WAR_ctonegateportlet&p_p_lifecycle=1&p_p_state=normal&p_p_mode=view&_DanhMuc_WAR_ctonegateportlet_javax.portlet.action=viewDanhMucThuTuc"
                    driver.get(url_dvc)
                        
                    #Lấy giá trị ô B3
                    tthc_lc = self.df_luanchuyen.iat[1, 0]  # hàng thứ 2 (A6) và cột thứ 1
                        
                    #click vào ô hiển thị tìm kiếm
                    dvc_click = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "//a[contains(text(), 'Hiển thị tìm kiếm')]")))
                    dvc_click.click()
                        
                    #click vào ô mã loại
                    ma_click = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.ID, "ma")))
                    ma_click.send_keys(tthc_lc)
                        
                    #click vào nút tìm kiếm
                    search_click = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "//button[@name='btnSearch' and @value='search']")))
                    search_click.click()
                    
                    # url_update_dvc = url + "/group/guest/danh-muc?p_p_id=DanhMuc_WAR_ctonegateportlet&p_p_lifecycle=1&p_p_state=normal&p_p_mode=view&p_p_col_id=column-2&p_p_col_count=1&_DanhMuc_WAR_ctonegateportlet_token=b95873f80bd19a49a85e927682122e1e&_DanhMuc_WAR_ctonegateportlet_thuTucId=145108&_DanhMuc_WAR_ctonegateportlet_javax.portlet.action=viewEditDanhMucThuTuc"
                    # driver.get(url_update_dvc)    

                    table1 = driver.find_element(By.CLASS_NAME, "table-data")
                    rows = table1.find_elements(By.TAG_NAME, "tr")

                    for row in rows:
                        try:
                            tenthutuc_cell = row.find_element(By.CLASS_NAME, "table-cell.tenthutuc")
                            link = tenthutuc_cell.find_element(By.TAG_NAME, "a")
                            link.click()
                        except StaleElementReferenceException:
                            # Xử lý nếu phát hiện lỗi StaleElementReferenceException
                            print("Element is no longer attached to the DOM. Retrying...")
                            continue

                    
                    
                           
                    #chọn quy trình gắn vào TTHC
                    chonqt_tthc = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CLASS_NAME, "select2-selection__rendered")))
                    chonqt_tthc.click()
                        
                    dien_tthc = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CLASS_NAME, "select2-search__field")))
                    dien_tthc.send_keys(tenqt_value)
                    dien_tthc.send_keys(Keys.ENTER) #Nhấn Enter để lưu
                    time.sleep(1)       
                    
                    #Lưu
                    save_tthc = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.ID, "luuKhongDongBo")))
                    save_tthc.click()
                     
                #Thao tac dong bo thu tuc chung
                if self.checkbox_dongboTTC.get():
                    print(f"Đang đồng bộ TTC cho hàng {index+1}")
                    driver.get(value3)
                    dongbo_ttc = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.ID, "btnDongBo")))
                    dongbo_ttc.click()
                    time.sleep(15)
                
                #Thao tac dong bo dich vu cong
                if self.checkbox_dongboDVC.get():
                    print(f"Đang đồng bộ DVC cho hàng {index+1}")
                    driver.get(value4)
                    dongbo_dvc = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.ID, "btnDongBo")))
                    dongbo_dvc.click()
                    # Tìm và nhấp vào nút có thuộc tính onclick
                    sync_button = WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.XPATH, '//a[@onclick="dongBoThuTuc();"]')))
                    sync_button.click()
                    time.sleep(33)
                    
                #Thao tac sao chép sổ văn bản
                if self.checkbox_copysovb_var.get():
                    print(f"Đang copy sổ văn bản cho hàng {index+1}")
                    driver.get(value7)
                    copy_sovb_from = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.ID, "_DanhMuc_WAR_ctonegateportlet_namCopy")))
                    copy_sovb_from.send_keys(now_year)
                    copy_sovb_to = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.ID, "_DanhMuc_WAR_ctonegateportlet_namMoi")))
                    copy_sovb_to.send_keys(next_year)
                    # Tìm và nhấp vào nút có thuộc tính onclick
                    copy_button = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.CLASS_NAME, "btn-primary")))
                    copy_button.click()
                    time.sleep(8)

                #Thao tác đồng bộ lĩnh vực
                if self.checkbox_dongbolv.get():
                    print(f"Đang đồng bộ Lĩnh vực cho hàng {index+1}")
                    driver.get(value6)
                    dongbo_linhvuc = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.ID, "btnDongBo")))
                    dongbo_linhvuc.click()
                    time.sleep(7)
                
                       
                # Thao tác cấu hình nghỉ lễ
                if self.checkbox_offnamsau_var.get():
                    
                    holiday_year = next_year if self.holiday_year_choice.get() == "Năm sau" else now_year
                    print(f"Đang cấu hình Ngày Nghỉ lễ cho hàng {index+1}")
                    driver.get(value5)
                    
                    # Cấu hình nghỉ Tết dương lịch
                    driver.get(value5)
                    tieude_duonglich = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.ID, "_quanlylichlamviec_WAR_ctonegatecoreportlet_eventName")))
                    
                    holiday_text = f"Nghỉ Tết Dương Lịch năm {holiday_year}"
                    tieude_duonglich.send_keys(holiday_text)
                    
                    thoigian_duonglich_from = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.ID, "_quanlylichlamviec_WAR_ctonegatecoreportlet_fromDateSchedule")))
                    thoigian_duonglich_from.send_keys(duonglich_from)
                    
                    thoigian_duonglich_to = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.ID, "_quanlylichlamviec_WAR_ctonegatecoreportlet_toDateSchedule")))
                    thoigian_duonglich_to.send_keys(duonglich_to)
                    
                    thoigian_duonglich_to.send_keys(Keys.ENTER)
                    thoigian_duonglich_to.send_keys(Keys.ENTER)
                    time.sleep(3)

                    # Cấu hình nghỉ Tết nguyên đáng
                    driver.get(value5)
                    tieude_nguyendan = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.ID, "_quanlylichlamviec_WAR_ctonegatecoreportlet_eventName")))
                    
                    nguyendan_text = f"Nghỉ Tết Nguyên Đán năm {holiday_year}"
                    tieude_nguyendan.send_keys(nguyendan_text)
                    
                    thoigian_nguyendan_from = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.ID, "_quanlylichlamviec_WAR_ctonegatecoreportlet_fromDateSchedule")))
                    thoigian_nguyendan_from.send_keys(nguyendan_from)
                    
                    thoigian_nguyendan_to = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.ID, "_quanlylichlamviec_WAR_ctonegatecoreportlet_toDateSchedule")))
                    thoigian_nguyendan_to.send_keys(nguyendan_to)
                    
                    thoigian_nguyendan_to.send_keys(Keys.ENTER)
                    thoigian_nguyendan_to.send_keys(Keys.ENTER)
                    time.sleep(3)


                    #Cấu hình Nghỉ Giỗ Tổ Hùng Vương
                    driver.get(value5)
                    tieude_gioto = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.ID, "_quanlylichlamviec_WAR_ctonegatecoreportlet_eventName")))
                    
                    gioto_text = f"Nghỉ Giỗ Tổ Hùng Vương 10/3 năm {holiday_year}"
                    tieude_gioto.send_keys(gioto_text)
                    
                    thoigian_gioto_from = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.ID, "_quanlylichlamviec_WAR_ctonegatecoreportlet_fromDateSchedule")))
                    thoigian_gioto_from.send_keys(gioto_from)
                    
                    thoigian_gioto_to = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.ID, "_quanlylichlamviec_WAR_ctonegatecoreportlet_toDateSchedule")))
                    thoigian_gioto_to.send_keys(gioto_to)
                    
                    thoigian_gioto_to.send_keys(Keys.ENTER)
                    thoigian_gioto_to.send_keys(Keys.ENTER)
                    time.sleep(3)

                    #Cấu hình nghỉ 30/4 và 1/5
                    driver.get(value5)
                    tieude_giaiphong = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.ID, "_quanlylichlamviec_WAR_ctonegatecoreportlet_eventName")))
                    
                    giaiphong_text = f"Nghỉ Lễ 30/4 và Quốc tế lao động 1/5 năm {holiday_year}"
                    tieude_giaiphong.send_keys(giaiphong_text)
                    
                    thoigian_giaiphong_from = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.ID, "_quanlylichlamviec_WAR_ctonegatecoreportlet_fromDateSchedule")))
                    thoigian_giaiphong_from.send_keys(giaiphong_from)
                    
                    thoigian_giaiphong_to = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.ID, "_quanlylichlamviec_WAR_ctonegatecoreportlet_toDateSchedule")))
                    thoigian_giaiphong_to.send_keys(giaiphong_to)
                    
                    thoigian_giaiphong_to.send_keys(Keys.ENTER)
                    thoigian_giaiphong_to.send_keys(Keys.ENTER)
                    time.sleep(3)

                    #Cấu hình nghỉ Lễ Quốc khánh 2/9
                    driver.get(value5)
                    tieude_quockhanh = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.ID, "_quanlylichlamviec_WAR_ctonegatecoreportlet_eventName")))
                    
                    quockhanh_text = f"Nghỉ Lễ Quốc Khánh 2/9 năm {holiday_year}"
                    tieude_quockhanh.send_keys(quockhanh_text)
                    
                    thoigian_quockhanh_from = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.ID, "_quanlylichlamviec_WAR_ctonegatecoreportlet_fromDateSchedule")))
                    thoigian_quockhanh_from.send_keys(quockhanh_from)
                    
                    thoigian_quockhanh_to = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.ID, "_quanlylichlamviec_WAR_ctonegatecoreportlet_toDateSchedule")))
                    thoigian_quockhanh_to.send_keys(quockhanh_to)
                    
                    thoigian_quockhanh_to.send_keys(Keys.ENTER)
                    thoigian_quockhanh_to.send_keys(Keys.ENTER)
                    time.sleep(3)

                logout_url = urljoin(driver.current_url, "/c/portal/logout")
                driver.get(logout_url)

                
        finally:
            driver.quit()
            self.stop_button.config(state=tk.DISABLED)
            print("ĐÃ HOÀN TẤT !!!")
            
    def open_quy_trinh_window(self):
        
        quy_trinh_window = tk.Toplevel(self.root)
        quy_trinh_window.title("CẤU HÌNH QUY TRÌNH")
        quy_trinh_window.geometry("1000x600")

        self.quy_trinh_data = []
        self.form_entries = []  # Khởi tạo form_entries
        self.luan_chuyen_entries = []  # Khởi tạo luan_chuyen_entries

        # Khung chứa Quy trình
        self.quy_trinh_frame = tk.LabelFrame(quy_trinh_window, text="Quy trình", font=self.default_font2)
        self.quy_trinh_frame.pack(fill="x", padx=10, pady=5)

        tk.Label(self.quy_trinh_frame, text="Tên quy trình:", font=self.default_font2).grid(row=0, column=0, padx=0, pady=0, sticky='w')
        self.ten_quy_trinh_entry = tk.Entry(self.quy_trinh_frame, font=self.default_font2, width=50)
        self.ten_quy_trinh_entry.grid(row=0, column=1, padx=0.5, pady=0.5)

        tk.Label(self.quy_trinh_frame, text="Bí danh quy trình:", font=self.default_font2).grid(row=1, column=0, padx=0, pady=0, sticky='w')
        self.bi_danh_entry = tk.Entry(self.quy_trinh_frame, font=self.default_font2, width=50)
        self.bi_danh_entry.grid(row=1, column=1, padx=0.5, pady=0.5)
        
        tk.Label(self.quy_trinh_frame, text="Gắn vào TTHC:", font=self.default_font2).grid(row=0, column=3, padx=5, pady=5, sticky='w')
        self.tthc_entry = tk.Entry(self.quy_trinh_frame, font=self.default_font2, width=50)
        self.tthc_entry.grid(row=0, column=4, padx=0.5, pady=0.5)
        
        # Khung chứa Danh sách Form
        self.danh_sach_form_frame = tk.LabelFrame(quy_trinh_window, text="Danh sách Form", font=self.default_font2)
        self.danh_sach_form_frame.pack(fill="x", padx=10, pady=5)

        # Thêm tiêu đề của các cột
        headers = ["ID", "Tên Form", "Mã Action", "Thời gian (Ngày)", "Nhóm người dùng", "Phòng ban"]
        for col, header in enumerate(headers):
            tk.Label(self.danh_sach_form_frame, text=header, font=self.default_font2).grid(row=0, column=col, padx=5, pady=5)
            


        self.add_form_entry()  # Thêm form entry đầu tiên

        self.add_form_button = tk.Button(self.danh_sach_form_frame, text="Thêm 1 hàng", command=self.add_form_entry, font=self.button_font2, fg="purple")
        self.add_form_button.grid(row=999, column=0, columnspan=5, pady=5)

        # Nút "Xong"
        self.done_button = tk.Button(self.danh_sach_form_frame, text="Xong", command=self.save_form_state, font=self.button_font2, fg="green")
        self.done_button.grid(row=999, column=4, columnspan=5, pady=5)

        # Khung chứa Danh sách Luân chuyển
        self.danh_sach_luan_chuyen_frame = tk.LabelFrame(quy_trinh_window, text="Danh sách Luân chuyển", font=self.default_font2)
        self.danh_sach_luan_chuyen_frame.pack(fill="x", padx=10, pady=5)

        self.add_luan_chuyen_entry()  # Thêm luân chuyển entry đầu tiên

        self.add_luan_chuyen_button = tk.Button(self.danh_sach_luan_chuyen_frame, text="Thêm 1 hàng", command=self.add_luan_chuyen_entry, font=self.button_font2, fg="purple")
        self.add_luan_chuyen_button.grid(row=999, column=0, columnspan=3, pady=5)

        # Nút tải về
        self.download_button = tk.Button(self.danh_sach_luan_chuyen_frame, text="Tải về", command=self.export_to_excel, font=self.button_font2, fg="blue")
        # self.download_button.pack(pady=5) # nút Tải về nằm rời bên dưới
        self.download_button.grid(row=999, column=3, columnspan=5, pady=5)
        
        


    def add_form_entry(self, id=None):
        if id is None:
            id = len(self.form_entries) + 1

        row = len(self.form_entries) + 1

        id_label = tk.Label(self.danh_sach_form_frame, text=str(id), font=self.default_font2)
        id_label.grid(row=row, column=0, padx=0.5, pady=0.5)

        ten_form_entry = tk.Entry(self.danh_sach_form_frame, font=self.default_font2, width=40)
        ten_form_entry.grid(row=row, column=1, padx=0.5, pady=0.5)

        action_menu = ttk.Combobox(self.danh_sach_form_frame, values=["Thêm mới", "Chuyển xử lý", "Trình phê duyệt", "Chuyển ban hành", "Chuyển trả kết quả"], font=self.default_font2, state='readonly')
        action_menu.grid(row=row, column=2, padx=0.5, pady=0.5)

        thoi_gian_entry = tk.Entry(self.danh_sach_form_frame, font=self.default_font2)
        thoi_gian_entry.grid(row=row, column=3, padx=0.5, pady=0.5)

        nhom_nguoi_dung_entry = tk.Entry(self.danh_sach_form_frame, font=self.default_font2)
        nhom_nguoi_dung_entry.grid(row=row, column=4, padx=0.5, pady=0.5)

        phongban_entry = tk.Entry(self.danh_sach_form_frame, font=self.default_font2)
        phongban_entry.grid(row=row, column=5, padx=0.5, pady=0.5)
        
        delete_button = tk.Button(self.danh_sach_form_frame, text="Xóa", command=lambda: self.delete_form_entry(id), font=self.button_font2, fg="red")
        delete_button.grid(row=row, column=6, padx=0.5, pady=0.5)

        self.form_entries.append((id_label, ten_form_entry, action_menu, thoi_gian_entry, nhom_nguoi_dung_entry, phongban_entry, delete_button))
        self.update_luan_chuyen_menus()

    def delete_form_entry(self, id):
        for entry in self.form_entries:
            if entry[0].cget("text") == str(id):
                for widget in entry:
                    widget.destroy()
                self.form_entries.remove(entry)
                break
        self.update_luan_chuyen_menus()
        self.reorder_form_entries()

    def reorder_form_entries(self):
        for idx, entry in enumerate(self.form_entries):
            entry[0].config(text=str(idx + 1))
            for widget in entry:
                widget.grid_configure(row=idx + 1)

    
    def add_luan_chuyen_entry(self):
        row = len(self.luan_chuyen_entries) + 1

        from_form_label = tk.Label(self.danh_sach_luan_chuyen_frame, text="Từ Form", font=self.default_font2)
        from_form_label.grid(row=row, column=0, padx=0.5, pady=0.5)

        from_form_menu = ttk.Combobox(self.danh_sach_luan_chuyen_frame, values=[form[1].get() for form in self.form_entries], font=self.default_font2, state='readonly')
        from_form_menu.grid(row=row, column=1, padx=0.5, pady=0.5)

        to_form_label = tk.Label(self.danh_sach_luan_chuyen_frame, text="Đến Form", font=self.default_font2)
        to_form_label.grid(row=row, column=2, padx=0.5, pady=0.5)

        to_form_menu = ttk.Combobox(self.danh_sach_luan_chuyen_frame, values=[form[1].get() for form in self.form_entries], font=self.default_font2, state='readonly')
        to_form_menu.grid(row=row, column=3, padx=0.5, pady=0.5)

        to_form_2_label = tk.Label(self.danh_sach_luan_chuyen_frame, text="Đến Form 2", font=self.default_font2)
        to_form_2_label.grid(row=row, column=4, padx=0.5, pady=0.5)

        to_form_2_menu = ttk.Combobox(self.danh_sach_luan_chuyen_frame, values=[form[1].get() for form in self.form_entries], font=self.default_font2, state='readonly')
        to_form_2_menu.grid(row=row, column=5, padx=0.5, pady=0.5)
        
        to_form_3_label = tk.Label(self.danh_sach_luan_chuyen_frame, text="Đến Form 3", font=self.default_font2)
        to_form_3_label.grid(row=row, column=6, padx=0.5, pady=0.5)

        to_form_3_menu = ttk.Combobox(self.danh_sach_luan_chuyen_frame, values=[form[1].get() for form in self.form_entries], font=self.default_font2, state='readonly')
        to_form_3_menu.grid(row=row, column=7, padx=0.5, pady=0.5)

        delete_button = tk.Button(self.danh_sach_luan_chuyen_frame, text="Xóa", command=lambda: self.delete_luan_chuyen_entry(row), font=self.button_font2, fg="red")
        delete_button.grid(row=row, column=8, padx=0.5, pady=0.5)

        self.luan_chuyen_entries.append((from_form_label, from_form_menu, to_form_label, to_form_menu, to_form_2_label, to_form_2_menu, to_form_3_label, to_form_3_menu, delete_button))

    def delete_luan_chuyen_entry(self, row):
        for entry in self.luan_chuyen_entries:
            if entry[0].grid_info()["row"] == row:
                for widget in entry:
                    widget.destroy()
                self.luan_chuyen_entries.remove(entry)
                break
        self.reorder_luan_chuyen_entries()

    def reorder_luan_chuyen_entries(self):
        for idx, entry in enumerate(self.luan_chuyen_entries):
            for widget in entry:
                widget.grid_configure(row=idx + 1)

    def update_luan_chuyen_menus(self):
        form_names = [form[1].get() for form in self.form_entries]
        for entry in self.luan_chuyen_entries:
            entry[1]['values'] = form_names
            entry[3]['values'] = form_names
            entry[5]['values'] = form_names
            entry[7]['values'] = form_names

    def save_form_state(self):
        self.update_luan_chuyen_menus()
        messagebox.showinfo("Thông báo", "Trạng thái của danh sách Form đã được lưu và cập nhật danh sách luân chuyển.")

    def export_to_excel(self):
        ten_quy_trinh = self.ten_quy_trinh_entry.get()
        bi_danh = self.bi_danh_entry.get() 
        tthc = self.tthc_entry.get() 
        
        quy_trinh_data = []
        for entry in self.form_entries:
            id = entry[0].cget("text")
            ten_form = entry[1].get()
            action = entry[2].get()
            thoi_gian = entry[3].get()
            nhom_nguoi_dung = entry[4].get()
            phongban = entry[5].get()
            quy_trinh_data.append([tthc,ten_quy_trinh, bi_danh, id, ten_form, action, thoi_gian, nhom_nguoi_dung, phongban ])

        luan_chuyen_data = []
        for entry in self.luan_chuyen_entries:
            from_form = entry[1].get()
            to_form = entry[3].get()
            to_form2 = entry[5].get()
            to_form3 = entry[7].get()
            luan_chuyen_data.append([tthc,ten_quy_trinh,from_form, to_form, to_form2, to_form3])

        df_quy_trinh = pd.DataFrame(quy_trinh_data, columns=["TTHC","Tên quy trình","Bí danh", "ID", "Tên Form", "Mã Action", "Thời gian", "Nhóm người dùng", "Phòng ban"])
        df_luan_chuyen = pd.DataFrame(luan_chuyen_data, columns=["TTHC","Tên quy trình", "Từ Form", "Đến Form", "Đến Form 2", "Đến Form 3"])

        wb = Workbook()
        ws_quy_trinh = wb.active
        ws_quy_trinh.title = "QuyTrinh"

        # Thêm tên quy trình và bí danh vào tiêu đề
        ws_quy_trinh.append(["Tên quy trình:", ten_quy_trinh])
        ws_quy_trinh.append(["Bí danh:", bi_danh])
        ws_quy_trinh.append(["Gắn vào TTHC:", tthc])
        ws_quy_trinh.append([])  # Thêm dòng trống để tách biệt

        for row in dataframe_to_rows(df_quy_trinh, index=False, header=True):
            ws_quy_trinh.append(row)
       
        for cell in ws_quy_trinh[5]:  
            cell.font = Font(bold=True)

        ws_luan_chuyen = wb.create_sheet(title="LuanChuyen")
        
        ws_luan_chuyen.append(["Tên quy trình:", ten_quy_trinh])
        ws_luan_chuyen.append(["Bí danh:", bi_danh])
        ws_luan_chuyen.append(["Gắn vào TTHC:", tthc])
        ws_luan_chuyen.append([])  # Thêm dòng trống để tách biệt
        
        for row in dataframe_to_rows(df_luan_chuyen, index=False, header=True):
            ws_luan_chuyen.append(row)

        for cell in ws_luan_chuyen[5]:
            cell.font = Font(bold=True)

        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            wb.save(file_path)
            messagebox.showinfo("Thông báo", f"File đã được lưu tại {file_path}")

    def attach_file(self):
        self.attached_file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if self.attached_file_path:
            self.attached_file_label.config(text=self.attached_file_path)
            try:
                self.df_quytrinh = pd.read_excel(self.attached_file_path, sheet_name='QuyTrinh', header=4)
                messagebox.showinfo("Thông báo", "Đã tải dữ liệu từ file quy trình.")
            except Exception as e:
                messagebox.showerror("Lỗi", f"Không thể tải dữ liệu từ file: {e}")
    
    def download_sample_file(self):
        # Specify the directory where the sample file is located
        sample_file_path = os.path.join(sys._MEIPASS, "data.xlsx")

        # Specify the directory where you want to save the downloaded file
        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if save_path:
            try:
                # Copy the sample file to the specified directory
                import shutil
                shutil.copy(sample_file_path, save_path)
                messagebox.showinfo("Thông báo", "Tải file mẫu thành công!")
            except Exception as e:
                messagebox.showerror("Lỗi", f"Lỗi khi tải file mẫu: {e}")
                


if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()
